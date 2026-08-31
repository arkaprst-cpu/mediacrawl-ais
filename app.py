"""
Media Crawl AIS — Pusat Strategi Kebijakan Pengawasan BPKP
Streamlit web app: input keyword → query expansion → crawl → analisis → download Excel
Provider AI: DeepSeek (nama model diatur lewat Secrets DEEPSEEK_MODEL,
default "deepseek-v4-flash" — lihat _baca_deepseek_model())
"""

import streamlit as st
import feedparser, json, time, re, io, threading
from datetime import datetime
from urllib.parse import quote_plus
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Klien DeepSeek (lazy import) ──
def get_deepseek_client(api_key: str):
    from openai import OpenAI
    return OpenAI(api_key=api_key, base_url="https://api.deepseek.com")


# ── Antrean crawl lintas-sesi ─────────────────────────────────────────────
# Latar belakang: dokumentasi resmi DeepSeek (api-docs.deepseek.com/
# quick_start/rate_limit) menyatakan limitnya berbasis CONCURRENCY per akun
# (500 koneksi API bersamaan untuk deepseek-v4-pro, 2.500 untuk
# deepseek-v4-flash) — bukan RPM/TPM seperti provider lain. Karena app ini
# memproses artikel satu-per-satu secara berurutan per sesi (bukan paralel),
# satu sesi crawl paling banter cuma punya 1 request DeepSeek yang sedang
# "in-flight" di satu waktu. Jadi puluhan pengguna bersamaan (mis. satu
# kelas) masih jauh di bawah limit DeepSeek itu sendiri — bukan itu yang
# butuh diamankan.
#
# Yang justru perlu diamankan: satu proses Streamlit yang dipakai bersama
# semua pengguna. Kalau banyak orang menekan "Mulai Crawl" nyaris
# bersamaan, semua request Google News RSS + parsing + panggilan DeepSeek
# + pembuatan Excel itu jalan di proses Python yang SAMA — bisa bikin
# semuanya kerasa lambat/berat kalau tidak dibatasi. _CrawlSlotManager di
# bawah ini singleton (di-cache lewat st.cache_resource, jadi SATU
# instance untuk SEMUA sesi/pengguna, bukan per-sesi) yang membatasi
# berapa banyak crawl boleh berjalan BERSAMAAN; sisanya otomatis antre
# dengan status yang jelas di layar, bukan dipaksa jalan sekaligus atau
# ditolak begitu saja.
class _CrawlSlotManager:
    def __init__(self, max_concurrent: int):
        self._lock = threading.Lock()
        self._active = 0
        self._max = max_concurrent

    def status(self):
        with self._lock:
            return self._active, self._max

    def acquire_blocking(self, on_wait=None, poll_seconds: float = 1.0):
        """Blok sampai dapat slot. on_wait(active, max) dipanggil tiap kali
        masih menunggu, supaya UI bisa menampilkan status antre real-time
        (Streamlit tetap mengirim update elemen ke browser meski script
        masih berjalan/nge-sleep di dalam loop ini — pola yang sama
        dipakai progress bar crawl di bawah)."""
        while True:
            with self._lock:
                if self._active < self._max:
                    self._active += 1
                    return
                active_sekarang = self._active
            if on_wait:
                on_wait(active_sekarang, self._max)
            time.sleep(poll_seconds)

    def release(self):
        with self._lock:
            self._active = max(0, self._active - 1)


def _baca_max_crawl_bersamaan(default: int = 5) -> int:
    """Bisa diubah tanpa edit kode lewat Streamlit Secrets
    (MAX_CRAWL_BERSAMAAN) — default 5 aman untuk server kelas biasa;
    naikkan kalau servernya cukup kuat, atau kalau ternyata tidak perlu
    seketat itu."""
    if not hasattr(st, "secrets"):
        return default
    try:
        return int(st.secrets.get("MAX_CRAWL_BERSAMAAN", default))
    except (TypeError, ValueError):
        return default


@st.cache_resource
def _get_crawl_slot_manager():
    return _CrawlSlotManager(_baca_max_crawl_bersamaan())


# ── Nama model DeepSeek — sengaja TIDAK di-hardcode di 3 tempat pemanggilan.
# DeepSeek pernah menghentikan nama model lama ("deepseek-chat" ->
# "deepseek-v4-flash", per 2026-07-24) tanpa jaminan itu tidak akan terulang
# untuk generasi berikutnya. Dengan nama model dibaca dari Secrets
# (DEEPSEEK_MODEL), kalau DeepSeek suatu saat mengganti/mem-pensiunkan
# "deepseek-v4-flash", cukup ubah satu nilai di Secrets — tidak perlu edit
# kode atau deploy ulang.
def _baca_deepseek_model(default: str = "deepseek-v4-flash") -> str:
    if not hasattr(st, "secrets"):
        return default
    nilai = st.secrets.get("DEEPSEEK_MODEL", default)
    return nilai if isinstance(nilai, str) and nilai.strip() else default


DEEPSEEK_MODEL = _baca_deepseek_model()


st.set_page_config(
    page_title="Media Crawl AIS — Pusat Strategi Kebijakan Pengawasan",
    page_icon="📰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Akses publik ke Repositori Isu (tanpa password) ─────────────────────
# Dicek sebelum gerbang password agar Repositori Isu bisa diakses via
# ?page=repositori tanpa login.
akses_publik_repositori = st.query_params.get("page") == "repositori"

if akses_publik_repositori:
    exec(open('repositori_isu.py').read())
    st.stop()

# ── Gerbang password (satu password untuk seluruh tim) ──────────────────
# Melindungi halaman Crawl & Analisis dan Dashboard AIS (keduanya dieksekusi
# lewat exec() di app.py, jadi satu gate ini cukup). Tidak berlaku untuk
# Repositori Isu — lihat akses_publik_repositori di atas.
def cek_password():
    if st.session_state.get("ais_authenticated", False):
        return True

    # Layar ini dieksekusi SEBELUM blok <style> font utama di bawah (yang
    # baru jalan setelah login) — tanpa import sendiri di sini, gerbang
    # password jatuh ke font default browser dan kelihatan polos/putus
    # nyambung dengan identitas amber-navy di halaman-halaman lain.
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@500;700&display=swap');
    html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
    .st-key-login_card {
        max-width: 440px; margin: 64px auto 0;
        background: linear-gradient(135deg, #0D1B2A 0%, #1C3D5A 100%);
        border-radius: 14px; border-top: 3px solid #F5A623;
        box-shadow: 0 12px 40px rgba(0,0,0,0.35);
        padding: 8px 8px 20px;
    }
    .login-kicker {
        font-family: 'IBM Plex Mono', monospace; font-size: 11px; font-weight: 700;
        letter-spacing: 0.18em; color: #F5A623; text-transform: uppercase; margin-bottom: 10px;
    }
    .login-title { font-size: 22px; font-weight: 700; color: #fff; line-height: 1.3; margin-bottom: 6px; }
    .login-org   { font-size: 12px; font-weight: 600; color: rgba(255,255,255,0.55); margin-bottom: 14px; }
    .login-desc  { font-size: 12.5px; color: rgba(255,255,255,0.68); line-height: 1.6; max-width: 320px; margin: 0 auto 22px; }
    </style>
    """, unsafe_allow_html=True)

    # Judul (nama lengkap) dan nama unit dipisah jadi dua level tipografi
    # sendiri-sendiri, bukan digabung satu baris dengan em dash — supaya
    # tidak ambigu mana judul aplikasi dan mana nama organisasinya. Ditambah
    # satu kalimat penjelasan singkat: sebelum ini gerbang password langsung
    # menyodorkan form login tanpa info apa pun soal aplikasinya sendiri.
    with st.container(key="login_card"):
        st.markdown("""
        <div style='text-align:center'>
          <div class="login-kicker">AIS</div>
          <div class="login-title">Analisis Isu Strategis Pengawasan</div>
          <div class="login-org">Pusat Strategi Kebijakan Pengawasan BPKP</div>
          <div class="login-desc">Pemantauan media otomatis untuk pengawasan isu strategis.</div>
        </div>
        """, unsafe_allow_html=True)

        pw_input = st.text_input("Password Akses Tim", type="password", key="pw_gate_input")
        masuk = st.button("Masuk", use_container_width=True, type="primary")

        if masuk:
            pw_benar = st.secrets.get("APP_PASSWORD", "") if hasattr(st, "secrets") else ""
            if not pw_benar:
                st.error("APP_PASSWORD belum dikonfigurasi di Streamlit Secrets. Hubungi pengelola aplikasi.")
            elif pw_input == pw_benar:
                st.session_state["ais_authenticated"] = True
                st.rerun()
            else:
                st.error("Password salah.")

        st.markdown("<div style='text-align:center;margin-top:16px;font-size:11px;opacity:0.5'>Mencari hasil analisis isu? <a href='?page=repositori' style='color:#F5A623'>Buka Repositori Isu Strategis publik →</a></div>", unsafe_allow_html=True)

    return False

if not cek_password():
    st.stop()

# ── NAVIGASI ───────────────────────────────────────────────────────────────
with st.sidebar:
    # CSS ini berlaku di semua halaman (blok with st.sidebar ini selalu
    # dieksekusi di setiap run, tidak seperti CSS spesifik per-halaman di
    # bawah) — merapatkan jarak antar elemen sidebar yang sebelumnya terlalu
    # lebar, plus styling tombol logout & label section yang mentereng.
    st.markdown("""
    <style>
    [data-testid="stSidebar"] hr { margin: 0.5rem 0; }
    [data-testid="stSidebar"] .stMarkdown { margin-bottom: 0; }
    .sidebar-section-label {
        font-size: 0.78rem; font-weight: 700; color: #F5A623;
        text-transform: uppercase; letter-spacing: 0.04em;
        margin: 2px 0 4px 0;
    }
    /* Tombol logout ditumpuk di pojok kanan-atas card brand (position:
       absolute) alih-alih dipepetkan ke kolom sempit — supaya lebar tombol
       menyesuaikan isi teksnya sendiri, tidak wrap/gepeng di layar sempit. */
    .st-key-sidebar_brand_row { position: relative; margin-bottom: 12px; }
    .st-key-logout_btn { position: absolute; top: 8px; right: 8px; z-index: 5; }
    .st-key-logout_btn button {
        background: transparent; border: 1px solid rgba(255,255,255,0.25);
        color: rgba(255,255,255,0.75); font-size: 11px; font-weight: 500;
        padding: 1px 10px; min-height: 24px; width: auto; white-space: nowrap;
    }
    .st-key-logout_btn button:hover {
        background: rgba(220,60,60,0.15); border-color: #dc3c3c; color: #ff9494;
    }
    /* Navigasi 3-menu — "tile" bukan radio polos, biar 3 fungsi utama app
       (crawl, klasterisasi/analisis, repositori) kelihatan sebagai 3 modul
       yang setara, bukan sekadar daftar link. Diskusi & di-tes langsung:
       st.radio TIDAK BISA dikasih subteks per-opsi (label-nya cuma teks
       polos 1 baris), jadi diganti 3x st.button, masing-masing labelnya
       pakai markdown 2 baris ("**Judul**\\n\\nSubteks kecil") — ini
       DIDUKUNG native oleh st.button (dites di probe terpisah), jauh lebih
       simpel & stabil daripada trik "tombol transparan ditumpuk di atas
       markdown" yang sempat dipertimbangkan. Highlight tile yang lagi aktif
       diatur lewat CSS dinamis di bawah (bukan di sini), karena butuh tahu
       st.session_state["ais_page"] saat itu.
    */
    /* Opsi 2 (revisi ruang sidebar): tile dipangkas jadi 1 baris (icon +
       judul saja, subteks dihapus dari sini) — subteksnya sudah nongol lagi
       persis sama di judul halaman utama, jadi di sidebar cukup labelnya
       biar nggak makan tinggi. Ini yang bikin field Kata Kunci Isu & Nama
       File nggak kegeser turun terlalu jauh. */
    [class*="st-key-navtile_"] { margin-bottom: 4px; }
    [class*="st-key-navtile_"] button {
        width: 100% !important; text-align: left !important;
        white-space: nowrap !important; height: auto !important;
        line-height: 1.4 !important; padding: 8px 14px !important;
        border-radius: 8px !important;
        background: rgba(255,255,255,0.04) !important;
        border: 1px solid rgba(255,255,255,0.12) !important;
    }
    [class*="st-key-navtile_"] button p { font-size: 13px !important; }
    </style>
    """, unsafe_allow_html=True)

    with st.container(key="sidebar_brand_row"):
        st.markdown("""
        <div style='background:linear-gradient(135deg,#0D1B2A,#1C3D5A);
                    border-radius:8px;padding:14px 16px;
                    border-bottom:2px solid #F5A623'>
          <div style='font-family:monospace;font-size:36px;font-weight:800;color:#F5A623;line-height:1.1'>AIS</div>
          <div style='font-size:10px;color:rgba(255,255,255,0.6);margin-top:4px'>Pusat Strategi Kebijakan Pengawasan BPKP</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Keluar", key="logout_btn", help="Keluar dari sesi"):
            st.session_state.pop("ais_authenticated", None)
            st.rerun()

    # Nama & subteks final hasil diskusi — subteksnya sengaja dibuat SAMA
    # PERSIS dengan subtitle di judul besar tiap halaman (lihat main-header
    # / ais-topbar di app.py, dashboard_ais.py, repositori_isu.py), supaya
    # nav sidebar & judul halaman "ngomong" hal yang sama dari 2 tempat.
    NAV_ITEMS = [
        {"id": "crawl",   "icon": "🔍", "label": "Crawl Berita",              "desc": "Tarik & analisis berita baru"},
        {"id": "klaster", "icon": "📊", "label": "Klasterisasi & Analisis", "desc": "Klasterisasi, tren & telaah"},
        {"id": "repo",    "icon": "🗄️", "label": "Repositori Isu Strategis",  "desc": "Arsip hasil yang sudah direview"},
    ]

    if "ais_page" not in st.session_state:
        st.session_state["ais_page"] = "crawl"

    # Highlight tile aktif — di-generate dinamis tiap rerun karena butuh
    # tahu halaman mana yang lagi aktif saat ini.
    st.markdown(f"""
    <style>
    .st-key-navtile_{st.session_state['ais_page']} button {{
        background: rgba(245,166,35,0.14) !important;
        border: 1px solid rgba(245,166,35,0.45) !important;
        border-left: 3px solid #F5A623 !important;
    }}
    .st-key-navtile_{st.session_state['ais_page']} button p:first-child {{ color: #F5A623 !important; }}
    </style>
    """, unsafe_allow_html=True)

    for item in NAV_ITEMS:
        with st.container(key=f"navtile_{item['id']}"):
            if st.button(
                f"{item['icon']} **{item['label']}**",
                key=f"navbtn_{item['id']}",
                use_container_width=True,
            ):
                if st.session_state["ais_page"] != item["id"]:
                    st.session_state["ais_page"] = item["id"]
                    st.rerun()

    page = st.session_state["ais_page"]
    st.divider()

# ── Excel builder (top-level agar bisa dipanggil dari dashboard_ais.py) ──
def buat_excel(data: list, label_isu: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Identifikasi Isu"

    C_NAVY="1F3864"; C_WHITE="FFFFFF"; C_SUB="D9E1F2"; C_ODD="EEF2F7"; C_EVEN="FFFFFF"
    TONE_C={"Positif":"C6EFCE","Netral":"FFEB9C","Negatif":"FFC7CE"}
    HEADERS=["No","Klaster Isu","Tanggal","Sumber","Link/Bukti","Judul/Post","Ringkasan Isu","Isu/Subisu","Aktor/Lokasi","Tone Berita","Risiko","Area Perhatian","Kondisi Klaster","Relevansi Pengawasan",
             "Sektor","Tema","Topik","Dampak/Implikasi (Final)","Gap Pengawasan","Usulan Pengawasan","Status Review",
             # Ditambahkan di paling AKHIR (kolom ke-22), SENGAJA bukan disisipkan
             # dekat "Relevansi Pengawasan" — supaya posisi kolom 1-21 (termasuk
             # blok telaah manusia Sektor..Status Review di 15-21) tidak bergeser
             # dan file lama (21 kolom) tetap terbaca benar oleh dashboard_ais.py.
             "Dimensi Pengawasan (GRCC AnCoDe)"]
    NCOL=len(HEADERS)
    COL_STATUS_REVIEW = HEADERS.index("Status Review") + 1  # posisi eksplisit,
    # bukan diasumsikan "kolom terakhir" — supaya penambahan kolom baru di masa
    # depan tidak diam-diam salah mewarnai kolom yang salah (ini yang dulu
    # nyaris kejadian pas nambah kolom Dimensi Pengawasan di atas).
    s=Side(style="thin",color="CCCCCC")
    BD=Border(left=s,right=s,top=s,bottom=s)

    def style(c,bg,bold=False,sz=9,center=False,fc=C_NAVY):
        c.font=Font(name="Arial",size=sz,bold=bold,color=fc)
        c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal="center" if center else "left",vertical="top",wrap_text=True)
        c.border=BD

    ws.merge_cells(f"A1:{get_column_letter(NCOL)}1")
    c=ws["A1"]; c.value="IDENTIFIKASI ISU HARIAN — ANALISIS ISU STRATEGIS PENGAWASAN"
    style(c,C_NAVY,bold=True,sz=13,center=True,fc=C_WHITE); ws.row_dimensions[1].height=28

    ws.merge_cells(f"A2:{get_column_letter(NCOL)}2")
    c=ws["A2"]
    c.value=f"Isu: {label_isu}  |  Generate: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  Total: {len(data)} artikel  |  Pusat Strategi Kebijakan Pengawasan BPKP"
    style(c,C_SUB,sz=9); ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=5

    for col,h in enumerate(HEADERS,1):
        c=ws.cell(row=4,column=col,value=h); style(c,C_NAVY,bold=True,sz=10,center=True,fc=C_WHITE)
    ws.row_dimensions[4].height=34

    for i,d in enumerate(data):
        r=5+i; bg=C_ODD if i%2==0 else C_EVEN
        baris=[i+1,d.get("klaster","-"),d.get("tanggal","-"),d.get("sumber","-"),d.get("link","-"),d.get("judul","-"),
               d.get("ringkasan_isu","-"),d.get("isu_subisu","-"),d.get("aktor_lokasi","-"),
               d.get("tone","Netral"),d.get("risiko","-"),d.get("area_perhatian","-"),
               d.get("kondisi_pemicu","-"),d.get("relevansi_pengawasan","-"),
               d.get("sektor","-"),d.get("tema","-"),d.get("topik","-"),
               d.get("dampak_implikasi_final","-"),d.get("gap_pengawasan","-"),d.get("usulan_pengawasan","-"),
               d.get("status_review","Belum Direview"),
               ", ".join(d.get("dimensi_pengawasan") or [])]
        for col,val in enumerate(baris,1):
            c=ws.cell(row=r,column=col,value=val); style(c,bg)
        tone_val=d.get("tone","Netral")
        ws.cell(row=r,column=10).fill=PatternFill("solid",fgColor=TONE_C.get(tone_val,"FFEB9C"))
        status_val=d.get("status_review","Belum Direview")
        STATUS_C={"Sudah Direview":"C6EFCE","Belum Direview":"F2F2F2"}
        ws.cell(row=r,column=COL_STATUS_REVIEW).fill=PatternFill("solid",fgColor=STATUS_C.get(status_val,"F2F2F2"))
        ws.row_dimensions[r].height=60

    for col,w in enumerate([5,28,12,18,35,40,45,25,25,12,45,40,45,40,30,28,40,45,40,45,16,40],1):
        ws.column_dimensions[get_column_letter(col)].width=w

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ══════════════════════════════════════════════════════════════════════════
# HALAMAN 1 — CRAWL & ANALISIS
# ══════════════════════════════════════════════════════════════════════════
if page == "crawl":

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');
    html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
    .main-header {
        background: linear-gradient(135deg, #1F3864 0%, #2d5299 100%);
        color: white; padding: 14px 20px; border-radius: 10px; margin-bottom: 14px;
    }
    .main-header h1 { font-size: 1.6rem; font-weight: 700; margin: 0 0 2px 0; }
    .main-header p  { font-size: 0.85rem; opacity: 0.75; margin: 0; font-family: 'IBM Plex Mono', monospace; }
    /* Kartu panduan "Cara memulai" — dibuat setara gaya kartu "Belum Ada
       Data" di Dashboard AIS (border putus-putus, background transparan
       tipis, warna teks ikut tema dark via `inherit`/opacity) supaya kedua
       halaman terasa satu sistem visual yang sama, bukan white-card ala
       dokumen yang kontras dengan tema dark aplikasi. */
    .st-key-crawl_empty_guide {
        border: 2px dashed rgba(128,128,128,0.3) !important;
        border-radius: 12px !important;
        background: rgba(128,128,128,0.06) !important;
        padding: 20px 24px !important;
        margin-bottom: 1.5rem !important;
    }
    .empty-guide-title { font-size: 16px; font-weight: 600; color: inherit; margin-bottom: 10px; text-align: center; }
    .empty-guide ol { margin: 0; padding-left: 1.2rem; max-width: 520px; margin-left: auto; margin-right: auto; }
    .empty-guide li { margin-bottom: 0.45rem; font-size: 12.5px; color: inherit; opacity: 0.75; line-height: 1.6; }
    .empty-guide li b { opacity: 1; font-weight: 600; }
    /* Kartu "Unduh Hasil" — sengaja beraksen amber (primaryColor app ini)
       supaya jadi satu-satunya elemen berwarna di antara kartu-kartu
       putih/netral lainnya, menandakan ini aksi penutup yang paling
       penting. Judul + tombol disatukan dalam satu container (bukan
       heading kecil terpisah dari tombol polos seperti sebelumnya) supaya
       bobot visualnya terasa setara. */
    .st-key-download_cta {
        background: linear-gradient(135deg, rgba(245,166,35,0.16), rgba(245,166,35,0.05)) !important;
        border: 1px solid rgba(245,166,35,0.4) !important;
        border-radius: 12px !important;
        padding: 22px 24px !important;
    }
    .download-cta-icon { font-size: 26px; text-align: center; margin-bottom: 8px; }
    .download-cta-title { font-size: 16px; font-weight: 700; color: #F5A623; margin-bottom: 4px; text-align: center; }
    .download-cta-sub { font-size: 12px; color: inherit; opacity: 0.7; margin-bottom: 16px; line-height: 1.5; text-align: center; }
    .stat-card {
        background: white; border: 1px solid #e8ecf0; border-radius: 10px;
        padding: 1.2rem 1.5rem; text-align: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    .stat-number { font-size: 2rem; font-weight: 700; color: #1F3864; line-height: 1; }
    .stat-label  { font-size: 0.75rem; color: #6b7280; margin-top: 0.4rem; text-transform: uppercase; letter-spacing: 0.05em; }
    .artikel-card {
        background: #ffffff; border: 1px solid #e2e8f0;
        border-left: 4px solid #1F3864; border-radius: 8px;
        padding: 1.2rem 1.5rem; margin-bottom: 1rem;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    }
    .artikel-judul { font-size: 1rem; font-weight: 600; color: #1e293b; margin-bottom: 0.4rem; line-height: 1.4; }
    /* Warna link judul dibuat permanen (bukan cuma muncul saat hover) —
       sebelumnya sama persis dengan warna teks biasa saat tidak disentuh
       kursor, jadi nyaris tidak kelihatan kalau judulnya bisa diklik.
       Panah "↗" juga diperkuat (opacity & ketebalan naik) sebagai sinyal
       kedua. Opsi ini dipilih karena "teks biru = bisa diklik" adalah
       konvensi web paling universal, tidak perlu penjelasan tambahan. */
    .artikel-judul a.judul-link { color: #1a56c4; text-decoration: none; }
    .artikel-judul a.judul-link:hover { color: #1F3864; text-decoration: underline; }
    .artikel-judul a.judul-link::after { content: "↗"; font-size: 0.8em; font-weight: 700; opacity: 0.75; margin-left: 5px; white-space: nowrap; }
    .artikel-meta  { font-size: 0.78rem; color: #64748b; margin-bottom: 0.8rem; font-family: 'IBM Plex Mono', monospace; }
    .artikel-ringkasan { font-size: 0.88rem; color: #374151; line-height: 1.6; margin-bottom: 0.6rem; }
    .tone-positif { background:#d1fae5;color:#065f46;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600; }
    .tone-netral  { background:#fef3c7;color:#92400e;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600; }
    .tone-negatif { background:#fee2e2;color:#991b1b;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600; }
    /* Pill sumber & badge aktor/topik — konsep sama dengan yang sudah
       dipakai di Dashboard AIS (pill nama media, badge topik indigo,
       badge aktor slate), tapi warnanya disesuaikan untuk card PUTIH di
       halaman ini (Dashboard AIS pakai card gelap) supaya kontrasnya
       tetap terjaga, bukan asal salin warna yang jadi pudar di background
       terang. */
    .pill-sumber {
        display: inline-block; font-size: 10px; font-weight: 700;
        font-family: 'IBM Plex Mono', monospace;
        color: #92400e; background: #fef3c7; border: 1px solid #fde68a;
        padding: 1px 7px; border-radius: 3px;
        margin-right: 6px; letter-spacing: 0.02em;
        text-transform: uppercase; vertical-align: middle;
    }
    .badge-pill {
        display: inline-block; font-size: 10.5px; font-weight: 600;
        font-family: 'IBM Plex Mono', monospace;
        padding: 2px 8px; border-radius: 4px;
        margin-right: 5px; margin-top: 4px;
    }
    .badge-topik { background: #e0e7ff; color: #4338ca; }
    .badge-aktor { background: #e2e8f0; color: #475569; }
    .artikel-tanggal { opacity: 0.65; }
    .query-box {
        background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 8px;
        padding: 0.8rem 1rem; margin-bottom: 1rem;
        font-size: 0.83rem; color: #1e40af; font-family: 'IBM Plex Mono', monospace;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Tier sumber ────────────────────────────────────────────────────────
    TIER1_KEYWORDS = {
        "kompas","tempo","detik","cnnindonesia","republika","antaranews",
        "mediaindonesia","bisnis","kontan","tribunnews","liputan6","okezone",
        "sindonews","jpnn","suara","kumparan","rmol","inews","katadata",
        "validnews","thejakartapost","jawapos",
    }

    def tier_sumber(url: str) -> str:
        domain = re.sub(r"https?://(www\.)?", "", url).split("/")[0].lower()
        for t1 in TIER1_KEYWORDS:
            if t1 in domain:
                return "Tier 1"
        return "Tier 2"

    def extract_sumber_dari_judul(judul: str) -> str:
        m = re.search(r"\s[-\u2013]\s([^-\u2013]+)$", judul.strip())
        if m:
            sumber = m.group(1).strip()
            sumber = re.sub(r"\s*\[.*?\]\s*$", "", sumber).strip()
            return sumber if sumber else ""
        return ""

    def bersihkan_judul_dari_sumber(judul: str) -> str:
        """Buang suffix ' - NamaSumber' Google News dari judul untuk
        ditampilkan \u2014 nama sumbernya sendiri sudah tampil terpisah sebagai
        pill (lihat pill-sumber), sama seperti tweak yang sudah diterapkan
        di Dashboard AIS."""
        m = re.search(r"\s[-\u2013]\s([^-\u2013]+)$", judul.strip())
        return judul[:m.start()].strip() if m else judul.strip()

    # ── PROMPT SISTEM (analisis per-artikel) ───────────────────────────────
    PROMPT_SISTEM = """Kamu adalah analis isu strategis pengawasan pemerintahan Indonesia untuk BPKP Pusat Strategi Kebijakan Pengawasan.

LANGKAH WAJIB — lakukan secara berurutan sebelum mengisi JSON:

LANGKAH 1 — BACA JUDUL DAN KONTEN SECARA LITERAL
Identifikasi: siapa yang disebut, apa yang terjadi, ada angka/besaran/tanggal apa, ada kata kunci negatif apa (dugaan, korupsi, gagal, mangkrak, turun, naik, dll). Kalau field "Konten" tersedia, gali detail konkret di dalamnya (nominal, jumlah, nama pihak/jabatan spesifik, tahapan/kegiatan yang disebutkan) — jangan cuma parafrase judul. Jangan tambahkan asumsi yang tidak ada di judul/konten.

LANGKAH 2 — ISI JSON
Output harus berupa JSON murni tanpa teks apapun di luar kurung kurawal:
{
  "ringkasan_isu"  : "3-4 kalimat FAKTUAL dan SEDETAIL MUNGKIN, gaya LEAD BERITA LANGSUNG (siapa-apa-kapan/di mana di kalimat pertama) — bukan esai yang mendeskripsikan artikelnya. WAJIB DIHINDARI, dalam kalimat mana pun (pembuka ATAUPUN penutup): frasa meta seperti 'Artikel [sumber] berjudul ...', 'Video berjudul ...', '... mengangkat isu/praktik/topik ...', '... membahas ...', '... mengulas ...', '... menyoroti ...', atau menyebut ulang nama sumber media dan tanggal publikasi (nama sumber & tanggal SUDAH tampil terpisah di UI, tidak perlu diulang di ringkasan). Langsung sampaikan FAKTANYA. Sebutkan nama program/institusi/jabatan spesifik, angka/nominal/tanggal kalau disebutkan, dan pemicu/konteksnya. Artikel ini harus bisa dipahami BERDIRI SENDIRI tanpa perlu membaca artikel lain di klasternya — utamakan detail konkret dari artikel ini sendiri, bukan kalimat generik yang bisa berlaku untuk artikel manapun di topik yang sama. JANGAN simpulkan risiko/relevansi/prioritas bagi pengawasan BPKP di sini — itu dinilai belakangan di level klaster, dengan konteks seluruh artikel sejenis, bukan per-artikel.

Contoh SALAH (meta-naratif, bertele-tele): 'Artikel Kompas.id berjudul Tak Lagi Nama, Ketika Calon Manajer Kopdes Merah Putih Dipanggil dengan Angka mengangkat praktik pemanggilan calon manajer Koperasi Desa (Kopdes) Merah Putih menggunakan angka, bukan nama. Hal ini terjadi dalam konteks seleksi calon manajer untuk program Kopdes Merah Putih. Artikel ini menyoroti aspek etika dan prosedur dalam proses seleksi manajer koperasi desa tersebut.'
Contoh BENAR (langsung ke fakta): 'Calon manajer Koperasi Desa (Kopdes) Merah Putih dipanggil menggunakan nomor urut, bukan nama, dalam proses seleksi program Kopdes Merah Putih yang digagas pemerintah. Praktik ini dinilai mengurangi harkat dan martabat calon manajer, yang seharusnya diperlakukan secara profesional dan manusiawi. Belum ada penjelasan resmi soal alasan penggunaan sistem penomoran ini dalam prosedur seleksi.'",
  "isu_subisu"     : "Nama isu utama / subisu spesifik (gunakan istilah dari judul, bukan abstraksi)",
  "aktor_lokasi"   : "Nama institusi atau jabatan yang disebut dalam judul / lokasi spesifik",
  "tone"           : "Positif" atau "Netral" atau "Negatif"
}

Aturan tambahan:
- Tone HANYA: Positif, Netral, atau Negatif
- Gunakan nama program/instansi/angka yang ada di judul/konten — jangan ganti dengan abstraksi
- Jika judul/konten tidak memberi cukup informasi, tetap isi semua field berdasarkan konteks topik crawl — jangan mengarang detail yang tidak ada sumbernya
- Bahasa Indonesia formal
- Output HANYA JSON murni"""

    # ── PROMPT KLASTER (klasterisasi + analisis risiko/area perhatian) ─────
    PROMPT_KLASTER = """Kamu adalah analis isu strategis pengawasan BPKP Pusat Strategi Kebijakan Pengawasan.

Kamu akan menerima daftar artikel (no, judul, ringkasan_isu, isu_subisu) hasil crawl SATU keyword/topik yang sama.
Meski semua artikel membahas topik yang sama, arah/akar persoalannya bisa berbeda-beda.

TUGAS 1 — KLASTERISASI:
Kelompokkan artikel-artikel ini ke dalam klaster isu utama berdasarkan KESAMAAN AKAR PERSOALAN DAN ARAH ISU — bukan sekadar kesamaan kata kunci permukaan.

ATURAN KLASTERISASI:
1. JUMLAH KLASTER 3 SAMPAI 5 — INI BATAS KERAS. Tidak boleh kurang dari 3, tidak boleh lebih dari 5, berapa pun banyaknya artikel — TERMASUK kalau seluruh artikel terasa membahas satu peristiwa/narasi besar yang sama.
2. Kalau artikel-artikelnya terasa homogen (satu topik/peristiwa besar yang sama), JANGAN jadikan itu alasan untuk menggabungkan semua jadi 1 klaster. Cari perbedaan SUDUT PANDANG, SKALA, atau FOKUS di antara artikel-artikel tersebut — misalnya: kasus/insiden spesifik vs tren atau daftar kumulatif yang berulang, fakta kejadian vs opini/analisis akar penyebab, level nasional vs level daerah/lokasi tertentu, atau tindakan pelaku vs respons institusi. Perbedaan semacam ini SELALU ada kalau dicari dengan teliti, dan itulah dasar pemecahan klasternya.
3. JANGAN membuat klaster terlalu granular berdasarkan detail permukaan (nama tokoh, nama acara, judul spesifik). Klaster harus berdasarkan AKAR PERSOALAN/ARAH ISU yang sama — kalau dua artikel sama-sama soal "kepercayaan investor" atau "tata kelola internal", gabungkan jadi satu klaster meski judul dan tokoh yang disebut berbeda.
4. Setiap artikel HARUS masuk tepat satu klaster (tidak ada yang terlewat, tidak ada duplikasi).

TUGAS 2 — ANALISIS RISIKO & AREA PERHATIAN PER KLASTER:
Untuk SETIAP klaster yang terbentuk, analisis berdasarkan KESELURUHAN artikel anggotanya (bukan satu artikel saja) untuk menjawab:
- "risiko": APA yang bisa terjadi jika kondisi/pola pada klaster ini tidak diintervensi (kerugian, kegagalan, penyimpangan). Ini PERNYATAAN RISIKO, bukan rencana kerja. JANGAN diawali label kategori (TATA KELOLA/PELAKSANAAN/KEBIJAKAN/EKSTERNAL) — tulis langsung isinya. DILARANG menulis frasa generik seperti "perlu transparansi", "perlu akuntabilitas", "tata kelola yang baik".
- "area_perhatian": TITIK LEMAH atau CELAH KONKRET yang melatarbelakangi risiko tersebut — bukan jenis kegiatan pengawasan yang harus dilakukan. JANGAN menulis "audit terhadap...", "reviu terhadap...", "perlu dilakukan pemeriksaan...". Tulis sebagai temuan/celah, bukan instruksi kerja.

Contoh SALAH (area_perhatian berbentuk kegiatan pengawasan):
"Audit kinerja dan keuangan terhadap pengelolaan dapur MBG oleh BGN"

Contoh BENAR (area_perhatian berbentuk titik lemah):
"Standar kebersihan dan kualitas bahan baku pada dapur penyedia MBG belum terverifikasi secara independen, sementara pengelolaan dapur melibatkan banyak penyedia pihak ketiga dengan pengawasan harian yang minim dari BGN"

TUGAS 3 — TAG DIMENSI PENGAWASAN (GRCC AnCoDe):
Tandai klaster ini dengan dimensi pengawasan mana saja yang BENAR-BENAR didukung isi artikelnya. Pilih HANYA dari daftar ini (boleh lebih dari satu, boleh KOSONG kalau memang tidak ada yang jelas-jelas cocok — JANGAN dipaksakan):
- "Governance": persoalan tata kelola/struktur kewenangan/akuntabilitas kelembagaan
- "Risk": ada indikasi risiko/ancaman terhadap capaian program yang belum terealisasi (bukan masalah yang sudah terjadi)
- "Control": ada indikasi lemahnya/tidak berjalannya mekanisme pengendalian internal
- "Compliance": ada indikasi ketidaksesuaian dengan aturan/regulasi/prosedur yang berlaku
- "Anti-Corruption": ada indikasi penyalahgunaan wewenang, gratifikasi, atau korupsi
- "Debottlenecking": ada indikasi hambatan birokrasi/regulasi yang memperlambat program prioritas nasional

Ini KLASIFIKASI, bukan narasi — jangan tulis penjelasan, cukup daftar nama dimensi yang cocok.

FORMAT OUTPUT setiap klaster wajib diberi:
- "nama": nama klaster singkat (maks 8 kata), mencerminkan isu utama bukan sekadar topik umum
- "kondisi_pemicu": 1-2 kalimat KONDISI/SITUASI konkret yang menyatukan artikel-artikel ini — NETRAL, sekadar menjelaskan latar/situasinya, TIDAK harus berupa "pemicu" atau kondisi negatif/kausal. Ikuti nada isi artikelnya apa adanya: kalau isunya soal upaya/program yang sedang berjalan (positif/netral), tulis apa adanya begitu — jangan dipaksa terdengar seperti ada yang salah. Framing risiko/kelemahan sepenuhnya jadi tugas "risiko" dan "area_perhatian" di bawah, bukan di sini.
- "risiko": sesuai aturan Tugas 2 di atas
- "area_perhatian": sesuai aturan Tugas 2 di atas
- "relevansi_pengawasan": mengapa klaster ini relevan/tidak terlalu prioritas bagi pengawasan BPKP
- "dimensi_pengawasan": array sesuai aturan Tugas 3 di atas (bisa array kosong [])
- "anggota": array berisi nomor (No) artikel yang masuk klaster ini

Urutkan array klaster dari yang paling kritikal/prioritas bagi pengawasan BPKP ke yang paling rendah prioritas.
Balas HANYA dalam format JSON murni, TANPA teks lain, TANPA markdown code fence, TANPA penjelasan di luar JSON.

Format output:
{"klaster": [{"nama": "...", "kondisi_pemicu": "...", "risiko": "...", "area_perhatian": "...", "relevansi_pengawasan": "...", "dimensi_pengawasan": ["Governance"], "anggota": [1,2,3]}]}
"""

    DIMENSI_PENGAWASAN_VALID = {"Governance", "Risk", "Control", "Compliance", "Anti-Corruption", "Debottlenecking"}

    # ── Query expansion ────────────────────────────────────────────────────
    def ekspansi_keyword_deepseek(client, keyword: str) -> list:
        prompt = f"""Kamu adalah asisten pencarian berita. Dari input keyword berikut, buat 4-5 variasi query pencarian berita yang lebih spesifik dan efektif untuk Google News.

Keyword input: "{keyword}"

Aturan:
- Variasikan dengan sinonim, singkatan, nama lokasi spesifik, atau aspek berbeda dari isu yang sama
- Gunakan bahasa Indonesia
- Kembalikan HANYA array JSON berisi string query, tanpa teks lain

Contoh output: ["query 1", "query 2", "query 3", "query 4"]"""
        try:
            resp = client.chat.completions.create(
                model=DEEPSEEK_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.4,
                max_tokens=300,
                # Matikan thinking mode: v4-flash aktifkan reasoning
                # chain-of-thought secara default (effort "high") yang jauh
                # lebih lambat dan tidak dibutuhkan untuk tugas sederhana
                # ini (query expansion). Tanpa ini, waktu respons bisa
                # berkali-kali lipat lebih lama dari perilaku deepseek-chat
                # lama, bahkan bisa menghabiskan max_tokens untuk proses
                # "berpikir" sebelum sempat menjawab.
                extra_body={"thinking": {"type": "disabled"}},
            )
            teks = resp.choices[0].message.content.strip()
            teks = re.sub(r"^```json\s*|^```\s*|\s*```$", "", teks).strip()
            queries = json.loads(teks)
            if isinstance(queries, list):
                return [keyword] + [q for q in queries if isinstance(q, str)]
        except Exception:
            pass
        return [keyword]

    # ── Filter video ───────────────────────────────────────────────────────
    DOMAIN_VIDEO = {"kompas.tv","metrotvnews.com","tvone.co.id","rctiplus.com","vidio.com","youtube.com","youtu.be"}
    JUDUL_VIDEO  = ["[full]","[live]","[video]","[breaking]","live streaming","siaran langsung","tonton video","nonton:","breaking news:","full video"]

    def is_video(judul: str, domain: str) -> bool:
        judul_lower = judul.lower()
        for dv in DOMAIN_VIDEO:
            if dv in domain.lower(): return True
        for marker in JUDUL_VIDEO:
            if marker in judul_lower: return True
        return False

    def bersihkan_snippet(snippet: str, judul: str) -> str:
        teks = snippet.replace("&nbsp;"," ").replace("&amp;","&")
        teks = re.sub(r"&[a-z]+;"," ",teks)
        teks = re.sub(r"<[^>]+>","",teks)
        teks = re.sub(r"\s+"," ",teks).strip()
        judul_bersih = re.sub(r"\s+"," ",judul).strip().lower()
        if teks.lower().startswith(judul_bersih[:40].lower()):
            return ""
        return teks

    # ── Crawl Google News RSS ──────────────────────────────────────────────
    def crawl_google_news(queries: list, max_articles: int) -> list:
        articles = []
        seen     = set()
        skipped  = 0
        headers  = {"User-Agent": "Mozilla/5.0 (compatible; AIS-Crawler/1.0)"}

        for q in queries:
            url = f"https://news.google.com/rss/search?q={quote_plus(q)}&hl=id&gl=ID&ceid=ID:id"
            try:
                feed = feedparser.parse(url, request_headers=headers)
                for entry in feed.entries:
                    link = entry.get("link","")
                    if link in seen: continue
                    seen.add(link)

                    judul  = entry.get("title","-")
                    domain = re.sub(r"https?://(www\.)?","",link).split("/")[0]
                    if is_video(judul, domain):
                        skipped += 1
                        continue

                    try:
                        tanggal = datetime(*entry.published_parsed[:3]).strftime("%d %b %Y")
                    except Exception:
                        pub = entry.get("published","")
                        tanggal = pub[:10] if pub else "-"

                    konten      = bersihkan_snippet(re.sub(r"<[^>]+>","",entry.get("summary","")), judul)
                    sumber_nama = extract_sumber_dari_judul(judul) or re.sub(r"https?://(www\.)?","",link).split("/")[0]
                    tier_asli   = tier_sumber(sumber_nama.lower())

                    articles.append({
                        "judul": judul, "link": link, "tanggal": tanggal,
                        "sumber": sumber_nama, "snippet": konten, "tier": tier_asli,
                    })
                    if len(articles) >= max_articles:
                        if skipped > 0:
                            st.caption(f"ℹ️ {skipped} artikel video dilewati.")
                        return articles
            except Exception as e:
                st.warning(f"Gagal crawl '{q}': {e}")

        if skipped > 0:
            st.caption(f"ℹ️ {skipped} artikel video dilewati.")
        return articles

    # ── Analisis: DeepSeek ─────────────────────────────────────────────────
    def analisis_deepseek(client, artikel: dict, rate_status=None) -> dict:
        konten = str(artikel.get("snippet","") or "").strip()
        konten_info = f"Konten  : {konten}" if konten else "Konten  : [tidak tersedia — analisis berdasarkan judul dan topik crawl]"
        prompt = (
            f"Topik crawl: {artikel.get('label_isu','-')}\n"
            f"Judul   : {artikel['judul']}\n"
            f"Sumber  : {artikel['sumber']}\n"
            f"Tanggal : {artikel['tanggal']}\n"
            f"{konten_info}\n\nHasilkan JSON analisis."
        )

        MAX_RETRY  = 4
        BASE_DELAY = 5

        for attempt in range(MAX_RETRY):
            try:
                resp = client.chat.completions.create(
                    model=DEEPSEEK_MODEL,
                    messages=[
                        {"role": "system", "content": PROMPT_SISTEM},
                        {"role": "user",   "content": prompt},
                    ],
                    temperature=0.3,
                    max_tokens=800,
                    # Matikan thinking mode — lihat catatan di
                    # ekspansi_keyword_deepseek(). Di sini dampaknya lebih
                    # kritis: 800 token bisa habis untuk "berpikir" sebelum
                    # model sempat menulis JSON hasil analisis, sehingga
                    # parsing gagal terus dan artikel jatuh ke fallback.
                    extra_body={"thinking": {"type": "disabled"}},
                )
                return _parse_json(resp.choices[0].message.content)

            except Exception as e:
                err_str = str(e).lower()
                if "429" in err_str or "rate" in err_str:
                    wait = BASE_DELAY * (2 ** attempt)
                    if rate_status:
                        rate_status.warning(f"⏳ Rate limit tercapai — menunggu {wait}s (retry {attempt+1}/{MAX_RETRY})...")
                    time.sleep(wait)
                elif attempt < MAX_RETRY - 1:
                    time.sleep(BASE_DELAY)
                    continue
                else:
                    return _fallback_error(artikel, str(e)[:200])

        return _fallback_error(artikel, "Rate limit tercapai — semua retry habis")

    # ── Klasterisasi + Analisis Risiko/Area Perhatian per klaster ──────────
    def klasterisasi_isu_deepseek(client, hasil_list: list) -> list:
        """Kirim seluruh ringkasan isu per-artikel ke DeepSeek untuk (1)
        dikelompokkan jadi 3-5 klaster isu utama, dan (2) dianalisis risiko
        & area perhatian berdasarkan keseluruhan artikel dalam tiap klaster
        (bukan per-artikel). Mengembalikan list klaster (bisa kosong jika
        gagal — pemanggil wajib menangani fallback)."""
        ringkasan_list = [
            {
                "no": i + 1,
                "judul": h.get("judul", "-"),
                "ringkasan_isu": h.get("ringkasan_isu", "-"),
                "isu_subisu": h.get("isu_subisu", "-"),
            }
            for i, h in enumerate(hasil_list)
        ]
        prompt = json.dumps(ringkasan_list, ensure_ascii=False)

        MAX_RETRY   = 4
        BASE_DELAY  = 6
        MIN_KLASTER = 3  # selaras dengan "batas keras" di PROMPT_KLASTER
        catatan_koreksi = ""

        for attempt in range(MAX_RETRY):
            try:
                resp = client.chat.completions.create(
                    model=DEEPSEEK_MODEL,
                    messages=[
                        {"role": "system", "content": PROMPT_KLASTER},
                        {"role": "user",   "content": prompt + catatan_koreksi},
                    ],
                    temperature=0.2,
                    max_tokens=3500,
                    # Matikan thinking mode — lihat catatan di
                    # ekspansi_keyword_deepseek(). Ini yang bikin tahap
                    # "Mengelompokkan isu & menganalisis risiko per
                    # klaster..." terasa muter lama: input klasterisasi
                    # berisi ringkasan SELURUH artikel sekaligus, jadi kalau
                    # thinking mode aktif, reasoning-nya jauh lebih panjang.
                    extra_body={"thinking": {"type": "disabled"}},
                )
                teks = (resp.choices[0].message.content or "").strip()

                if not teks:
                    wait = BASE_DELAY * (2 ** attempt)
                    time.sleep(wait)
                    continue

                teks = re.sub(r"^```json\s*|^```\s*|\s*```$", "", teks).strip()
                m = re.search(r"\{.*\}", teks, flags=re.DOTALL)
                if m:
                    teks = m.group(0)
                parsed = json.loads(teks)
                klaster_list = parsed.get("klaster", [])

                # Validasi: pastikan setiap anggota klaster adalah int yang valid
                total_artikel = len(hasil_list)
                anggota_terpakai = set()
                klaster_valid = []
                for kl in klaster_list:
                    anggota = [a for a in kl.get("anggota", []) if isinstance(a, int) and 1 <= a <= total_artikel]
                    if not anggota:
                        continue
                    anggota_terpakai.update(anggota)
                    # Sanitasi dimensi_pengawasan: cuma terima nilai dari daftar
                    # resmi (jaga-jaga AI berhalusinasi nama dimensi lain/typo),
                    # dan cuma terima kalau memang berbentuk list — bukan
                    # dipaksa selalu ada isinya (klaster tanpa dimensi jelas =
                    # list kosong, itu valid).
                    dimensi_mentah = kl.get("dimensi_pengawasan", [])
                    dimensi = [d for d in dimensi_mentah if isinstance(d, str) and d in DIMENSI_PENGAWASAN_VALID] if isinstance(dimensi_mentah, list) else []
                    klaster_valid.append({**kl, "anggota": anggota, "dimensi_pengawasan": dimensi})

                # Tegakkan batas BAWAH (min 3 klaster) di kode, bukan cuma di
                # prompt. PROMPT_KLASTER sudah menulis aturan ini sebagai
                # "batas keras", tapi teks prompt saja tidak selalu ditaati:
                # kalau kumpulan artikelnya sangat homogen (satu keyword besar
                # yang menarik banyak berita dari satu narasi yang sama, mis.
                # "OTT KPK" -> 25 artikel yang semuanya soal maraknya OTT
                # kepala daerah 2026), model bisa memutuskan itu cukup 1
                # klaster raksasa — padahal biasanya tetap ada sudut/skala
                # berbeda (kasus spesifik vs tren umum, fakta vs analisis akar
                # sebab, dst.) yang layak dipisah untuk kebutuhan pengawasan.
                # Kalau ini kejadian, retry dengan teguran eksplisit alih-alih
                # diam-diam meloloskan hasil yang melanggar aturannya sendiri
                # (analog dengan penegakan batas ATAS/maks 5 klaster di bawah).
                if len(klaster_valid) < MIN_KLASTER and attempt < MAX_RETRY - 1:
                    catatan_koreksi = (
                        f"\n\nPERINGATAN: Percobaan sebelumnya cuma menghasilkan "
                        f"{len(klaster_valid)} klaster untuk {total_artikel} artikel — "
                        f"ini MELANGGAR aturan wajib (JUMLAH KLASTER 3 SAMPAI 5, berapa "
                        f"pun banyaknya artikel, bahkan kalau akar masalahnya terasa "
                        f"sama). WAJIB pecah lebih detail kali ini: cari sudut pandang, "
                        f"skala, atau fokus yang berbeda dalam kumpulan artikel ini — "
                        f"misalnya kasus/insiden spesifik vs tren atau daftar kumulatif "
                        f"yang berulang, fakta kejadian vs opini/analisis akar penyebab, "
                        f"atau level nasional vs level daerah/lokasi tertentu — lalu "
                        f"jadikan masing-masing sudut sebagai klaster terpisah."
                    )
                    time.sleep(BASE_DELAY)
                    continue

                # Artikel yang tidak masuk klaster manapun -> klaster "Isu Lainnya"
                sisa = [n for n in range(1, total_artikel + 1) if n not in anggota_terpakai]
                if sisa:
                    klaster_valid.append({
                        "nama": "Isu Lainnya",
                        "kondisi_pemicu": "Artikel dengan arah isu yang tidak terkelompok ke klaster utama.",
                        "risiko": "-",
                        "area_perhatian": "-",
                        "relevansi_pengawasan": "Perlu ditelaah manual — tidak teridentifikasi pola yang jelas.",
                        "dimensi_pengawasan": [],
                        "anggota": sisa,
                    })

                # Tegakkan batas maks 5 klaster di kode (bukan hanya di prompt)
                MAKS_KLASTER = 5
                if len(klaster_valid) > MAKS_KLASTER:
                    klaster_valid.sort(key=lambda k: len(k.get("anggota", [])), reverse=True)
                    dipertahankan = klaster_valid[:MAKS_KLASTER - 1]
                    digabung = klaster_valid[MAKS_KLASTER - 1:]
                    anggota_gabungan = sorted(set(a for kl in digabung for a in kl.get("anggota", [])))
                    dipertahankan.append({
                        "nama": "Isu Lainnya",
                        "kondisi_pemicu": "Gabungan beberapa isu kecil yang tidak cukup signifikan untuk jadi klaster tersendiri.",
                        "risiko": "-",
                        "area_perhatian": "Perlu ditelaah manual per artikel — masing-masing berdiri sendiri tanpa pola dominan.",
                        "relevansi_pengawasan": "-",
                        "dimensi_pengawasan": [],
                        "anggota": anggota_gabungan,
                    })
                    klaster_valid = dipertahankan

                nama_lainnya = [kl for kl in klaster_valid if kl.get("nama") == "Isu Lainnya"]
                if len(nama_lainnya) > 1:
                    anggota_gab = sorted(set(a for kl in nama_lainnya for a in kl.get("anggota", [])))
                    klaster_valid = [kl for kl in klaster_valid if kl.get("nama") != "Isu Lainnya"]
                    klaster_valid.append({**nama_lainnya[0], "anggota": anggota_gab})

                return klaster_valid

            except Exception:
                if attempt < MAX_RETRY - 1:
                    time.sleep(BASE_DELAY)
                    continue
                return []  # fallback: dashboard/Excel akan tampil tanpa klaster

        return []

    # ── Helpers parse & fallback ───────────────────────────────────────────
    def _parse_json(teks: str) -> dict:
        teks = teks.strip()
        teks = re.sub(r"^```json\s*|^```\s*|\s*```$", "", teks).strip()
        m = re.search(r"\{.*\}", teks, flags=re.DOTALL)
        if m:
            teks = m.group(0)
        hasil = json.loads(teks)
        if hasil.get("tone") not in ("Positif", "Netral", "Negatif"):
            hasil["tone"] = "Netral"
        return hasil

    def _fallback_error(artikel: dict, pesan: str) -> dict:
        return {
            "ringkasan_isu": artikel.get("judul","-"),
            "isu_subisu": "-", "aktor_lokasi": "-",
            "tone": "Netral",
            "_error": pesan,
        }

    # ══════════════════════════════════════════════════════════════════════
    # SIDEBAR — konfigurasi provider & input
    # ══════════════════════════════════════════════════════════════════════
    with st.sidebar:
        # Header "Media Crawl AIS" dihapus — sudah terwakili oleh brand "AIS"
        # di nav atas, jadi ini dulu redundant dan makan tempat vertikal.

        # ── API Key DeepSeek ─────────────────────────────────────────────
        # Kalau sudah dikonfigurasi lewat Secrets, tidak perlu ditunjukkan
        # ke user — itu detail teknis provider AI, bukan sesuatu yang perlu
        # mereka pahami/pikirkan. UI cukup langsung siap pakai; kotak isian
        # API Key hanya muncul (plus divider-nya) kalau memang belum
        # dikonfigurasi — supaya tidak ada divider ganda yang mengapit ruang
        # kosong saat key sudah tersedia dari Secrets.
        deepseek_key_default = st.secrets.get("DEEPSEEK_API_KEY","") if hasattr(st,"secrets") else ""
        if deepseek_key_default:
            active_key = deepseek_key_default
        else:
            active_key = st.text_input("DeepSeek API Key", type="password", placeholder="sk-...")
            st.divider()

        # Label section custom (bukan label bawaan Streamlit) supaya field
        # kunci ini — kata kunci pencarian & nama file — terlihat mentereng
        # dan jelas, bukan sekadar label abu-abu standar.
        st.markdown('<div class="sidebar-section-label">🔍 Kata Kunci Isu</div>', unsafe_allow_html=True)
        keywords_raw = st.text_area(
            "Kata Kunci Isu",
            placeholder="Contoh:\nMBG, makan bergizi gratis\nDanantara ekspor\nPertamax BBM",
            height=120,
            label_visibility="collapsed",
        )
        st.markdown('<div class="sidebar-section-label">📁 Label Isu (nama file Excel)</div>', unsafe_allow_html=True)
        label_isu = st.text_input(
            "Label Isu (nama file Excel)",
            placeholder="Contoh: Pertamax BBM Juni 2026",
            label_visibility="collapsed",
        )
        max_art   = st.slider(
            "Maks. Artikel", min_value=5, max_value=25, value=20, step=5,
        )
        st.divider()
        # disabled=True selama crawl_running True — supaya tombol ini betul-
        # betul tidak bisa diklik lagi (bukan cuma "diabaikan") selama
        # proses crawl+analisis AI sedang berjalan. Lihat komentar di trigger
        # "if run_btn and not ... crawl_running" di bawah untuk alasan
        # lengkap kenapa ini butuh 1x rerun ekstra sebelum proses berat
        # mulai.
        run_btn = st.button(
            "🔍 Mulai Crawl", use_container_width=True,
            disabled=st.session_state.get("crawl_running", False),
        )

    # ── Main area ──────────────────────────────────────────────────────────
    # Provider AI (DeepSeek) sengaja tidak lagi ditampilkan sebagai pill di
    # header — sama seperti badge API key di sidebar, itu detail teknis
    # implementasi yang tidak perlu diketahui/dipikirkan user.
    st.markdown("""
    <div class="main-header">
      <h1>🔍 Crawl Berita</h1>
      <p>Tarik & analisis berita baru</p>
    </div>
    """, unsafe_allow_html=True)

    # Pesan hasil crawl sebelumnya (sukses/gagal/warning) — dicatat ke
    # session_state alih-alih langsung st.success()/st.error() di tempat
    # kejadian, karena alur proteksi klik-ganda di bawah SELALU diakhiri
    # st.rerun() supaya tombol "Mulai Crawl" kembali aktif; pesan yang
    # ditampilkan langsung sebelum rerun akan hilang sebelum sempat
    # terbaca. Jadi ditulis dulu ke sini, baru ditampilkan di render
    # berikutnya (persis setelah rerun).
    for _level, _pesan in st.session_state.pop("_crawl_pesan", []):
        getattr(st, _level)(_pesan)

    # Panduan awal — hanya tampil sebelum sesi ini pernah punya hasil crawl.
    # Begitu "hasil" masuk ke session_state (crawl pertama berhasil), guard
    # ini otomatis False untuk sisa sesi, jadi tidak mengganggu tampilan
    # hasil di rerun berikutnya. `not run_btn` mencegah kartu ini sempat
    # nongol sekilas di pass rerun saat tombol baru saja diklik. Judul
    # sengaja "Cara memulai" (bukan "Mulai di sini") karena kartu ini ada
    # di area konten utama, sedangkan aksi sesungguhnya (isi kata kunci,
    # klik tombol) ada di sidebar kiri — bukan di kartu ini sendiri.
    if "hasil" not in st.session_state and not run_btn:
        with st.container(key="crawl_empty_guide"):
            st.markdown("""
            <div class="empty-guide-title">👈 Cara memulai</div>
            <ol>
              <li>Isi <b>kata kunci isu</b> di sidebar kiri (boleh lebih dari satu, satu per baris)</li>
              <li>Beri <b>nama file Excel</b> yang mudah dikenali nanti</li>
              <li>Klik <b>🔍 Mulai Crawl</b> — hasil analisisnya akan muncul di halaman ini</li>
            </ol>
            """, unsafe_allow_html=True)

    # ── Trigger crawl ──────────────────────────────────────────────────────
    # Klik pertama HANYA menyalakan flag lalu langsung st.rerun() — belum
    # melakukan proses berat apa pun. Ini disengaja: tombol di sidebar
    # sudah terlanjur ter-render "aktif" ke browser SEBELUM baris ini
    # sempat dieksekusi, dan Streamlit tidak mengirim ulang status tombol
    # di tengah eksekusi skrip yang sama. Kalau proses crawl+analisis AI
    # (yang bisa makan waktu lama) langsung dijalankan di sini, tombolnya
    # akan tetap terlihat aktif & bisa diklik lagi selama proses
    # berlangsung — dan klik kedua itu BUKAN cuma diabaikan, tapi
    # membatalkan proses yang sedang jalan lalu memulai crawl baru dari
    # nol (buang kuota AI + waktu percuma). Rerun di sini memaksa render
    # ulang SEBELUM proses berat dimulai, supaya tombol terkirim ke
    # browser dalam kondisi disabled=True lebih dulu.
    if run_btn and not st.session_state.get("crawl_running", False):
        st.session_state["crawl_running"] = True
        st.rerun()

    if st.session_state.get("crawl_running", False):
        def _catat_pesan(pesan, level="info"):
            st.session_state.setdefault("_crawl_pesan", []).append((level, pesan))

        def _gagal(pesan, level="error"):
            _catat_pesan(pesan, level)
            st.rerun()

        try:
            if not active_key:
                _gagal("Masukkan API Key terlebih dahulu.")
            if not keywords_raw.strip():
                _gagal("Masukkan minimal satu kata kunci.")
            if not label_isu.strip():
                _gagal("Isi Label Isu untuk nama file Excel.")

            keywords_input = [k.strip() for k in re.split(r"[\n,]+", keywords_raw) if k.strip()]

            ai_client = get_deepseek_client(active_key)

            # ── Antrean crawl lintas-sesi ────────────────────────────────────
            # Kalau slot sedang penuh (banyak orang crawl bersamaan), tunggu di
            # sini dengan status yang jelas alih-alih membiarkan semua sesi
            # membebani proses Streamlit sekaligus. Lihat _CrawlSlotManager di
            # bagian atas file untuk alasan lengkapnya.
            slot_mgr = _get_crawl_slot_manager()
            antre_status = st.empty()
            def _lapor_antre(active, mx):
                antre_status.warning(
                    f"🕐 Sedang antre ({active}/{mx} slot terpakai). Permintaan Anda akan "
                    f"diproses otomatis begitu slot tersedia — mohon tunggu di halaman ini, "
                    f"jangan ditutup atau di-refresh."
                )
            slot_mgr.acquire_blocking(on_wait=_lapor_antre)
            antre_status.empty()

            try:
                st.subheader("⏳ Proses Crawl & Analisis")

                with st.spinner("Memperluas keyword..."):
                    all_queries = []
                    for kw in keywords_input:
                        expanded = ekspansi_keyword_deepseek(ai_client, kw)
                        all_queries.extend(expanded)

                query_lines = "<br>".join(f"🔍 {q}" for q in all_queries)
                st.markdown(f'<div class="query-box"><b>Query ({len(all_queries)} variasi):</b><br>{query_lines}</div>', unsafe_allow_html=True)

                prog_bar  = st.progress(0, text="Crawling Google News...")
                status_tx = st.empty()
                status_tx.info(f"Crawling {len(all_queries)} query...")
                artikel_raw = crawl_google_news(all_queries, max_art)

                if not artikel_raw:
                    _gagal("Tidak ada artikel ditemukan.", level="warning")

                status_tx.success(f"✅ {len(artikel_raw)} artikel ditemukan. Memulai analisis...")

                hasil_list  = []
                rate_status = st.empty()

                for idx, art in enumerate(artikel_raw):
                    pct = int((idx + 1) / len(artikel_raw) * 100)
                    prog_bar.progress(pct, text=f"Menganalisis artikel {idx+1}/{len(artikel_raw)}...")

                    # DeepSeek berbayar, tanpa RPM ketat -> delay ringan cukup
                    time.sleep(0.3)

                    art["label_isu"] = label_isu.strip()
                    analisis = analisis_deepseek(ai_client, art, rate_status)
                    hasil_list.append({**art, **analisis})

                rate_status.empty()
                prog_bar.progress(100, text="✅ Mengelompokkan jadi klaster isu...")

                with st.spinner("Mengelompokkan isu & menganalisis risiko per klaster..."):
                    klaster_list = klasterisasi_isu_deepseek(ai_client, hasil_list)

                # Sebarkan nama klaster, risiko, area_perhatian, kondisi_pemicu, dan
                # relevansi_pengawasan dari hasil klasterisasi ke setiap artikel
                # anggotanya — semua field ini TIDAK lagi dianalisis per-artikel,
                # melainkan diwarisi dari analisis tingkat klaster (lebih kaya
                # konteks, lebih sedikit panggilan AI).
                klaster_per_no = {}
                for kl in klaster_list:
                    for no in kl.get("anggota", []):
                        klaster_per_no[no] = kl
                for i, h in enumerate(hasil_list):
                    kl = klaster_per_no.get(i + 1)
                    if kl:
                        h["klaster"]               = kl.get("nama", "-")
                        h["risiko"]                = kl.get("risiko", "-")
                        h["area_perhatian"]        = kl.get("area_perhatian", "-")
                        h["kondisi_pemicu"]        = kl.get("kondisi_pemicu", "-")
                        h["relevansi_pengawasan"]  = kl.get("relevansi_pengawasan", "-")
                        h["dimensi_pengawasan"]    = kl.get("dimensi_pengawasan", [])
                    else:
                        h["klaster"]               = "-"
                        h["risiko"]                = "-"
                        h["area_perhatian"]        = "-"
                        h["kondisi_pemicu"]        = "-"
                        h["relevansi_pengawasan"]  = "-"
                        h["dimensi_pengawasan"]    = []

                prog_bar.progress(100, text="✅ Selesai!")
                status_tx.empty()

                st.session_state["hasil"]     = hasil_list
                st.session_state["klaster"]   = klaster_list
                st.session_state["label_isu"] = label_isu.strip()
                st.session_state["ais_ready"] = True
                st.session_state["ais_errors"] = [h.get("_error") for h in hasil_list if h.get("_error")]
                # Tandai crawl BARU ini sebagai sumber data TERAKTIF di Dashboard AIS
                # — sebelumnya Dashboard AIS selalu mengutamakan Excel upload manual
                # apa pun yang terjadi belakangan, jadi hasil crawl baru bisa
                # "kalah" ditimpa tampilan upload lama yang masih tersimpan di
                # sesi. Lihat dashboard_ais.py bagian "LOAD & PROCESS DATA".
                st.session_state["_dash_last_source"] = "session"

                if not klaster_list:
                    _catat_pesan("⚠️ Klasterisasi gagal — Excel & dashboard tetap tersedia, tapi tanpa pengelompokan isu, risiko, dan area perhatian.", level="warning")
                _catat_pesan("✅ Analisis selesai. Buka **📊 Klasterisasi & Analisis** di sidebar untuk visualisasi lengkap.", level="success")
            finally:
                # WAJIB dilepas apa pun yang terjadi (termasuk kegagalan
                # validasi/crawl di atas, atau error tak terduga) — kalau
                # slot bocor/tidak pernah dilepas, app ini akan makin
                # "penuh" terus-menerus sampai di-restart, dan pengguna
                # berikutnya antre selamanya walau sebenarnya tidak ada
                # crawl lain yang benar-benar berjalan.
                slot_mgr.release()
        finally:
            # Tombol "Mulai Crawl" dikunci lagi ke kondisi bisa diklik, apa
            # pun hasil akhirnya (sukses, gagal validasi via _gagal(), atau
            # error tak terduga) — supaya user tidak pernah macet dengan
            # tombol yang ke-disable permanen.
            st.session_state["crawl_running"] = False
        # Rerun terakhir supaya browser benar-benar menampilkan tombol
        # dalam keadaan aktif kembali (nilai session_state saja tidak
        # otomatis terkirim ulang ke frontend tanpa render baru), dan pesan
        # di _crawl_pesan tampil bersih di render berikutnya.
        st.rerun()

    # ── Error diagnostik ───────────────────────────────────────────────────
    if st.session_state.get("ais_errors"):
        errs  = st.session_state["ais_errors"]
        total = len(st.session_state.get("hasil", []))
        with st.expander(f"⚠️ {len(errs)} dari {total} artikel gagal dianalisis", expanded=True):
            st.code(errs[0])
            low = errs[0].lower()
            if "rate" in low or "429" in low or "quota" in low:
                st.warning("🕐 Rate limit — retry otomatis sudah berjalan. Jika masih banyak yang kosong, tunggu 1–2 menit lalu ulangi.")
            elif "auth" in low or "401" in low:
                st.warning("🔑 Masalah API Key. Cek kembali key di Streamlit Secrets.")
            elif "json" in low or "expecting" in low:
                st.warning("📋 Respons non-JSON dari AI. Biasanya sementara — coba ulangi.")

    # ── Hasil & download ───────────────────────────────────────────────────
    if "hasil" in st.session_state:
        hasil_list = st.session_state["hasil"]
        label_isu  = st.session_state["label_isu"]

        tone_counts = {"Positif":0,"Netral":0,"Negatif":0}
        for h in hasil_list:
            t = h.get("tone","Netral")
            tone_counts[t] = tone_counts.get(t,0) + 1

        c0,c1,c2,c3 = st.columns(4)
        with c0:
            st.markdown(f'<div class="stat-card"><div class="stat-number">{len(hasil_list)}</div><div class="stat-label">Total Artikel</div></div>', unsafe_allow_html=True)
        for col,tone,warna,emoji in [(c1,"Positif","#065f46","🟢"),(c2,"Netral","#92400e","🟡"),(c3,"Negatif","#991b1b","🔴")]:
            with col:
                st.markdown(f'<div class="stat-card"><div class="stat-number" style="color:{warna}">{tone_counts[tone]}</div><div class="stat-label">{emoji} {tone}</div></div>', unsafe_allow_html=True)

        # "Unduh Hasil" sengaja dipindah ke PALING BAWAH (setelah daftar
        # artikel, lihat blok "download-cta" di akhir) — sebelumnya
        # ditaruh di sini, tepat setelah statistik, malah memotong alur
        # sebelum user sempat lihat artikel apa saja yang ketarik.
        # Sekarang jadi penutup alami: tinjau dulu artikelnya, baru unduh.
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### 📰 Artikel yang Ditemukan")
        c_filter1, c_filter2 = st.columns(2)
        with c_filter1:
            filter_tone = st.selectbox("Filter tone:", ["Semua","Positif","Netral","Negatif"])
        with c_filter2:
            nama_klaster_list = ["Semua"] + sorted(set(h.get("klaster","-") for h in hasil_list))
            filter_klaster = st.selectbox("Filter klaster:", nama_klaster_list)

        tampil = hasil_list
        if filter_tone != "Semua":
            tampil = [h for h in tampil if h.get("tone") == filter_tone]
        if filter_klaster != "Semua":
            tampil = [h for h in tampil if h.get("klaster") == filter_klaster]

        # Caption ini sengaja eksplisit bilang "artikel individu" dan "tiap
        # kartu di bawah adalah satu artikel" — sebelumnya heading-nya
        # "Pratinjau Hasil Analisis" yang kesannya satu hasil analisis
        # gabungan, padahal isinya daftar artikel satu-satu.
        st.caption(f"Menampilkan **{len(tampil)} dari {len(hasil_list)}** artikel individu hasil crawl — tiap kartu di bawah adalah satu artikel.")

        for h in tampil:
            tone = h.get("tone","Netral")
            klaster_label = h.get("klaster","-")
            link = h.get("link","")
            judul_bersih = bersihkan_judul_dari_sumber(h.get("judul","-"))
            sumber_pill = f'<span class="pill-sumber">{h.get("sumber","-")}</span>' if h.get("sumber") else ""
            # Judul langsung jadi hyperlink ke artikel asli (bukan link
            # terpisah di bawah card via expander) — lebih cepat diakses,
            # dan tetap terasa satu kesatuan dengan card-nya. Nama sumber
            # dipisah dari teks judul dan ditampilkan sebagai pill di
            # depannya — sama seperti tweak pill sumber di Dashboard AIS.
            if link and link != "-":
                judul_html = f'{sumber_pill}<a href="{link}" target="_blank" rel="noopener" class="judul-link">{judul_bersih}</a>'
            else:
                judul_html = f'{sumber_pill}{judul_bersih}'

            # Aktor/Lokasi dipecah jadi pill per-nama (bukan satu blob teks),
            # dan Isu/Subisu ditampilkan sebagai badge topik — sama seperti
            # tweak badge topik vs. aktor yang sudah diterapkan di panel
            # Detail Analisis Dashboard AIS (dua level informasi berbeda,
            # jadi dipisah warna: topik indigo, aktor slate).
            daftar_aktor = [a.strip() for a in str(h.get("aktor_lokasi","")).split(",") if a.strip() and a.strip() != "-"]
            aktor_pills = "".join(f'<span class="badge-pill badge-aktor">👤 {a}</span>' for a in daftar_aktor)
            isu_subisu = h.get("isu_subisu","-")
            topik_badge = f'<span class="badge-pill badge-topik">🏷️ {isu_subisu}</span>' if isu_subisu and isu_subisu != "-" else ""

            st.markdown(f"""
            <div class="artikel-card">
                <div class="artikel-judul">{judul_html}</div>
                <div class="artikel-meta"><span class="artikel-tanggal">📅 {h.get('tanggal','-')}</span> &nbsp;·&nbsp; {h.get('tier','')} &nbsp;·&nbsp; <span class="tone-{tone.lower()}">{tone}</span></div>
                <div style="margin:2px 0 10px 0">{topik_badge}{aktor_pills}</div>
                <div class="artikel-ringkasan">{h.get('ringkasan_isu','-')}</div>
                <div style="margin-top:8px;font-size:0.75rem;color:#94a3b8">🗂️ Klaster: {klaster_label}</div>
            </div>""", unsafe_allow_html=True)

        # Kartu unduh — penutup alami setelah user selesai meninjau daftar
        # artikel di atas. Sengaja dibuat menonjol (aksen amber, warna
        # primaryColor app ini) supaya bobot visualnya setara dengan
        # pentingnya aksi ini — sebelumnya cuma heading kecil + tombol
        # polos, gampang terlewat padahal ini tujuan akhir dari crawl.
        st.markdown("<br>", unsafe_allow_html=True)
        with st.container(key="download_cta"):
            st.markdown("""
            <div class="download-cta-icon">📥</div>
            <div class="download-cta-title">Unduh Hasil Lengkap</div>
            <div class="download-cta-sub">Sudah selesai meninjau? Unduh seluruh artikel beserta analisis klaster, risiko, dan area perhatian dalam satu file Excel.</div>
            """, unsafe_allow_html=True)
            excel_buf = buat_excel(hasil_list, label_isu)
            nama_file = f"MediaCrawl_AIS_{label_isu.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button("📥 Download Excel", data=excel_buf, file_name=nama_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary")


# ══════════════════════════════════════════════════════════════════════════
# HALAMAN 2 — DASHBOARD AIS
# ══════════════════════════════════════════════════════════════════════════
elif page == "klaster":
    exec(open('dashboard_ais.py').read())

# ══════════════════════════════════════════════════════════════════════════
# HALAMAN 3 — REPOSITORI ISU
# ══════════════════════════════════════════════════════════════════════════
elif page == "repo":
    exec(open('repositori_isu.py').read())
